import json
import re
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from django.db import transaction
from openpyxl import load_workbook

from apps.ktru.models import (
    KtruCharacteristic,
    KtruCharacteristicValue,
    KtruGroup,
    KtruParseLog,
    KtruPosition,
    KtruPositionRelation,
)

EXCEL_GROUP_ALIASES = {
    "системный блок": ("системный блок", "Объем оперативной памяти"),
    "монитор": ("монитор", "Диагональ"),
    "мониторы": ("монитор", "Диагональ"),
    "клавиатура": ("клавиатура", "Тип подключения"),
    "мышь": ("мышь", "Тип мыши / тип подключения"),
    "моноблок": ("моноблок", "Диагональ"),
    "ноутбук": ("ноутбук", "Диагональ"),
}

HTML_GROUP_ALIASES = {
    "клавиатура": "клавиатура",
    "мониторы": "монитор",
    "монитор": "монитор",
    "моноблок": "моноблок",
    "мышь": "мышь",
    "ноутбук": "ноутбук",
    "системный блок": "системный блок",
}

REFINE_CHARACTERISTIC_ALIASES = {
    "системный блок": ["объем установленной оперативной памяти", "объем оперативной памяти"],
    "монитор": ["размер диагонали", "диагональ"],
    "клавиатура": ["тип подключения"],
    "мышь": ["тип подключения", "тип мыши"],
    "моноблок": ["размер диагонали", "диагональ"],
    "ноутбук": ["размер диагонали экрана", "диагональ"],
}


def normalize_code(code: str) -> str:
    return str(code).strip().replace(" ", "")


def _clean(value: str) -> str:
    return " ".join((value or "").split())


def _format_excel_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _clean(str(value))


def _excel_group_from_header(value: str) -> tuple[str, str] | None:
    text = _clean(str(value)).lower()
    if "номер ктру" not in text:
        return None
    for alias, group_data in EXCEL_GROUP_ALIASES.items():
        if alias in text:
            return group_data
    return None


def _is_excel_code_header(value) -> bool:
    return value is not None and "номер ктру" in str(value).lower()


def _read_text(path: Path) -> str:
    for encoding in ("utf-8", "cp1251"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="replace")


def load_seed_excel(path: Path) -> list[dict]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    result = []
    for row_index in range(1, sheet.max_row + 1):
        for column_index in range(1, sheet.max_column + 1):
            group_data = _excel_group_from_header(sheet.cell(row_index, column_index).value)
            if not group_data:
                continue
            group_name, refine_attribute_name = group_data
            for data_row in range(row_index + 1, sheet.max_row + 1):
                code = sheet.cell(data_row, column_index).value
                if _is_excel_code_header(code):
                    break
                refine_value = sheet.cell(data_row, column_index + 1).value
                if not code or not refine_value:
                    continue
                code = normalize_code(_format_excel_value(code))
                refine_value = _format_excel_value(refine_value)
                if not code or not refine_value:
                    continue
                result.append(
                    {
                        "group": group_name,
                        "refine_attribute_name": refine_attribute_name,
                        "code": code,
                        "refine_value": refine_value,
                    }
                )
    unique = {}
    for entry in result:
        unique[(entry["group"], entry["code"], entry["refine_value"])] = entry
    return list(unique.values())


def parse_ktru_print_form(url: str) -> dict:
    response = requests.get(url, timeout=20)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")
    text = soup.get_text("\n", strip=True)
    return {"raw_text": text, "raw_html": response.text, "source_url": url}


def _clear_position_characteristics(position: KtruPosition) -> None:
    characteristics = position.characteristics.all()
    if not characteristics.exists():
        return
    # Reseeding changes catalog rows. Existing specs keep display snapshots,
    # so remove stale links before replacing source characteristics.
    from apps.specs.models import TechnicalSpecItemCharacteristic

    TechnicalSpecItemCharacteristic.objects.filter(ktru_characteristic__in=characteristics).delete()
    characteristics.delete()


def _table_rows(table) -> list[list[str]]:
    rows = []
    for tr in table.find_all("tr"):
        cells = [_clean(cell.get_text(" ", strip=True)) for cell in tr.find_all(["td", "th"])]
        if any(cells):
            rows.append(cells)
    return rows


def _find_characteristics_table(soup: BeautifulSoup):
    for table in soup.find_all("table"):
        text = _clean(table.get_text(" ", strip=True))
        if "Наименование характеристики" in text and "Значение характеристики" in text:
            return table
    return None


def _metadata_from_html(soup: BeautifulSoup) -> dict:
    metadata = {}
    first_table = soup.find("table")
    if not first_table:
        return metadata
    for row in _table_rows(first_table):
        if len(row) >= 2 and row[0]:
            metadata[row[0]] = row[1]
    return metadata


def parse_eis_html_file(path: Path) -> dict:
    soup = BeautifulSoup(_read_text(path), "html.parser")
    metadata = _metadata_from_html(soup)
    table = _find_characteristics_table(soup)
    characteristics = []
    current = None
    if table:
        for row in _table_rows(table)[1:]:
            if len(row) >= 4:
                current = {
                    "name": row[0],
                    "kind": row[1],
                    "is_required": "Неизменяемая" in row[1],
                    "is_multiple_choice": "множественный" in row[1].lower(),
                    "unit_name": row[3],
                    "values": [],
                    "raw_data": {"kind": row[1]},
                }
                if row[2]:
                    current["values"].append(row[2])
                characteristics.append(current)
            elif len(row) >= 2 and current:
                if row[0]:
                    current["values"].append(row[0])
                if row[1] and not current["unit_name"]:
                    current["unit_name"] = row[1]

    return {
        "metadata": metadata,
        "characteristics": characteristics,
        "raw_text": soup.get_text("\n", strip=True),
    }


def _group_name_from_html(path: Path, parsed: dict) -> str | None:
    stem = path.stem.strip().lower()
    if stem in HTML_GROUP_ALIASES:
        return HTML_GROUP_ALIASES[stem]
    product_name = parsed.get("metadata", {}).get("Наименование товара, работы, услуги", "").lower()
    for alias, group in HTML_GROUP_ALIASES.items():
        if alias in product_name:
            return group
    return None


def _value_matches_refine(value: str, refine_value: str) -> bool:
    left = _clean(value).lower().replace(",", ".")
    right = _clean(refine_value).lower().replace(",", ".")
    if left == right:
        return True
    numbers = re.findall(r"\d+(?:\.\d+)?", left)
    return bool(numbers and right in numbers)


def _is_refine_characteristic(group_name: str, characteristic_name: str) -> bool:
    name = characteristic_name.lower()
    aliases = REFINE_CHARACTERISTIC_ALIASES.get(group_name, [])
    return any(alias in name for alias in aliases)


def apply_eis_html_to_group(group_name: str, parsed: dict) -> None:
    group = KtruGroup.objects.filter(name__iexact=group_name).first()
    if not group:
        return
    metadata = parsed.get("metadata", {})
    html_characteristics = parsed.get("characteristics", [])
    positions = KtruPosition.objects.filter(group=group, is_refined=True)
    for position in positions:
        position.name = metadata.get("Наименование товара, работы, услуги") or position.name
        position.okpd2_code = metadata.get("Код по ОКПД2") or position.okpd2_code
        position.okpd2_name = metadata.get("Наименование товара, работы, услуги по ОКПД2") or position.okpd2_name
        position.unit_name = (
            metadata.get("Единицы измерения (количество товара, объем работ, услуги по ОКЕИ)")
            or metadata.get("Единицы измерения(количество товара, объем работ, услуги по ОКЕИ)")
            or position.unit_name
        )
        position.raw_data = {**position.raw_data, "eis_html_metadata": metadata}
        position.parse_status = KtruPosition.ParseStatus.SUCCESS
        position.save()

        _clear_position_characteristics(position)
        for order, char_data in enumerate(html_characteristics, start=1):
            values = list(dict.fromkeys(str(value) for value in char_data.get("values", []) if str(value).strip()))
            if _is_refine_characteristic(group.name, char_data["name"]):
                matched = [value for value in values if _value_matches_refine(value, position.refine_value)]
                if matched:
                    values = matched
                elif position.refine_value:
                    values = [position.refine_value]
                is_required = True
            else:
                is_required = bool(char_data.get("is_required", False))
            characteristic = KtruCharacteristic.objects.create(
                ktru_position=position,
                name=char_data["name"],
                is_required=is_required,
                is_multiple_choice=char_data.get("is_multiple_choice", False),
                unit_name=char_data.get("unit_name", ""),
                instruction="",
                value_type="string",
                display_order=order,
                group_title="",
                raw_data=char_data.get("raw_data", {}),
            )
            for value_order, value in enumerate(values, start=1):
                KtruCharacteristicValue.objects.create(
                    characteristic=characteristic,
                    value=value,
                    display_order=value_order,
                    raw_data={"source": "eis_html"},
                )


def load_eis_html_dir(path: Path) -> int:
    if not path.exists():
        return 0
    loaded = 0
    for html_path in path.glob("*.html"):
        parsed = parse_eis_html_file(html_path)
        group_name = _group_name_from_html(html_path, parsed)
        if not group_name:
            KtruParseLog.objects.create(source="eis_html", status="failed", message=f"Cannot map {html_path.name}")
            continue
        apply_eis_html_to_group(group_name, parsed)
        KtruParseLog.objects.create(
            source="eis_html",
            status="success",
            message=f"Loaded {html_path.name} for group {group_name}: {len(parsed.get('characteristics', []))} characteristics",
            raw_response=parsed.get("raw_text", "")[:20000],
        )
        loaded += 1
    return loaded


@transaction.atomic
def upsert_excel_seed_entries(entries: list[dict]) -> None:
    for entry in entries:
        group, _ = KtruGroup.objects.update_or_create(
            name=entry["group"],
            defaults={"refine_attribute_name": entry["refine_attribute_name"]},
        )
        enlarged = group.positions.filter(is_enlarged=True).first()
        if not enlarged:
            enlarged, _ = KtruPosition.objects.update_or_create(
                code=f"ENLARGED-{group.name}",
                defaults={
                    "group": group,
                    "original_code_from_seed": f"ENLARGED-{group.name}",
                    "normalized_code": normalize_code(f"ENLARGED-{group.name}"),
                    "name": group.name,
                    "unit_name": "Штука",
                    "unit_code": "796",
                    "is_enlarged": True,
                    "is_refined": False,
                    "parse_status": KtruPosition.ParseStatus.FIXTURE,
                    "raw_data": {"source": "excel"},
                },
            )

        position, _ = KtruPosition.objects.update_or_create(
            code=entry["code"],
            defaults={
                "group": group,
                "original_code_from_seed": entry["code"],
                "normalized_code": normalize_code(entry["code"]),
                "name": enlarged.name or group.name,
                "okpd2_code": enlarged.okpd2_code,
                "okpd2_name": enlarged.okpd2_name,
                "unit_name": enlarged.unit_name or "Штука",
                "unit_code": enlarged.unit_code or "796",
                "is_enlarged": False,
                "is_refined": True,
                "parse_status": KtruPosition.ParseStatus.PENDING,
                "refine_value": entry["refine_value"],
                "raw_data": {"source": "excel", **entry},
            },
        )
        KtruPositionRelation.objects.update_or_create(
            enlarged_position=enlarged,
            refined_position=position,
            refine_attribute_value=entry["refine_value"],
            defaults={"refine_attribute_name": group.refine_attribute_name},
        )


@transaction.atomic
def upsert_fixture_data(data: dict) -> None:
    for group_data in data.get("groups", []):
        group, _ = KtruGroup.objects.update_or_create(
            name=group_data["name"],
            defaults={"refine_attribute_name": group_data["refine_attribute_name"]},
        )
        enlarged, _ = KtruPosition.objects.update_or_create(
            code=group_data["enlarged_code"],
            defaults={
                "group": group,
                "original_code_from_seed": group_data["enlarged_code"],
                "normalized_code": normalize_code(group_data["enlarged_code"]),
                "name": group_data["name"],
                "okpd2_code": group_data.get("okpd2_code", ""),
                "okpd2_name": group_data.get("okpd2_name", ""),
                "unit_name": group_data.get("unit_name", "шт."),
                "unit_code": group_data.get("unit_code", "796"),
                "is_enlarged": True,
                "is_refined": False,
                "parse_status": KtruPosition.ParseStatus.FIXTURE,
                "raw_data": group_data,
            },
        )
        for position_data in group_data.get("positions", []):
            position, _ = KtruPosition.objects.update_or_create(
                code=position_data["code"],
                defaults={
                    "group": group,
                    "original_code_from_seed": position_data["code"],
                    "normalized_code": normalize_code(position_data["code"]),
                    "name": position_data.get("name", group_data["name"]),
                    "okpd2_code": position_data.get("okpd2_code", group_data.get("okpd2_code", "")),
                    "okpd2_name": position_data.get("okpd2_name", group_data.get("okpd2_name", "")),
                    "unit_name": position_data.get("unit_name", group_data.get("unit_name", "шт.")),
                    "unit_code": position_data.get("unit_code", group_data.get("unit_code", "796")),
                    "description": position_data.get("description", ""),
                    "is_enlarged": False,
                    "is_refined": True,
                    "parse_status": KtruPosition.ParseStatus.FIXTURE,
                    "refine_value": str(position_data["refine_value"]),
                    "raw_data": position_data,
                },
            )
            KtruPositionRelation.objects.update_or_create(
                enlarged_position=enlarged,
                refined_position=position,
                refine_attribute_value=str(position_data["refine_value"]),
                defaults={"refine_attribute_name": group.refine_attribute_name},
            )
            _clear_position_characteristics(position)
            for order, char_data in enumerate(position_data.get("characteristics", []), start=1):
                characteristic = KtruCharacteristic.objects.create(
                    ktru_position=position,
                    name=char_data["name"],
                    is_required=char_data.get("is_required", False),
                    is_multiple_choice=char_data.get("is_multiple_choice", False),
                    unit_name=char_data.get("unit_name", ""),
                    instruction=char_data.get("instruction", ""),
                    value_type=char_data.get("value_type", "string"),
                    display_order=char_data.get("display_order", order),
                    group_title=char_data.get("group_title", ""),
                    raw_data=char_data,
                )
                for value_order, value in enumerate(char_data.get("values", []), start=1):
                    KtruCharacteristicValue.objects.create(
                        characteristic=characteristic,
                        value=str(value),
                        display_order=value_order,
                        raw_data={"value": value},
                    )


def load_fixture(path: Path) -> None:
    data = json.loads(path.read_text(encoding="utf-8"))
    upsert_fixture_data(data)
    KtruParseLog.objects.create(source="fixture", status="success", message=f"Loaded {path.name}")
